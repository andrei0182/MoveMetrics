from io import BytesIO

import openpyxl
import pandas as pd

from src.monthly_report.excel_generator import generate_monthly_workbook


def test_one_sheet_per_calendar_day():
    df = pd.DataFrame({
        "Job #": ["A", "B", "C"],
        "Date": ["2026-08-01", "2026-08-01", "2026-08-02"],
    })
    wb = openpyxl.load_workbook(BytesIO(generate_monthly_workbook(df)))

    assert "01.AUG" in wb.sheetnames
    assert "02.AUG" in wb.sheetnames


def test_day_sheet_contains_only_that_days_rows():
    df = pd.DataFrame({
        "Job #": ["A", "B", "C"],
        "Date": ["2026-08-01", "2026-08-01", "2026-08-02"],
    })
    wb = openpyxl.load_workbook(BytesIO(generate_monthly_workbook(df)))

    assert wb["01.AUG"].max_row - 1 == 2  # minus header
    assert wb["02.AUG"].max_row - 1 == 1


def test_total_sheet_contains_every_row():
    df = pd.DataFrame({
        "Job #": ["A", "B", "C"],
        "Date": ["2026-08-01", "2026-08-01", "2026-08-02"],
    })
    wb = openpyxl.load_workbook(BytesIO(generate_monthly_workbook(df)))

    assert wb["TOTAL"].max_row - 1 == 3


def test_missing_date_goes_to_unknown_date_sheet():
    df = pd.DataFrame({"Job #": ["A", "B"], "Date": ["2026-08-01", None]})
    wb = openpyxl.load_workbook(BytesIO(generate_monthly_workbook(df)))

    assert "Unknown Date" in wb.sheetnames
    assert wb["Unknown Date"].max_row - 1 == 1


def test_no_date_column_still_produces_total_sheet():
    df = pd.DataFrame({"Job #": ["A", "B"], "Charged": [100, 200]})
    wb = openpyxl.load_workbook(BytesIO(generate_monthly_workbook(df)))

    assert wb.sheetnames == ["TOTAL"]
    assert wb["TOTAL"].max_row - 1 == 2


def test_empty_dataframe_does_not_crash():
    df = pd.DataFrame({"Job #": [], "Date": []})
    wb = openpyxl.load_workbook(BytesIO(generate_monthly_workbook(df)))

    assert "TOTAL" in wb.sheetnames
